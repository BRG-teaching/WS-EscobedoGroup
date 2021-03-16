import os
import compas

from openpyxl import Workbook
from openpyxl import utils
from openpyxl.styles import Font

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, '../data')
FILE_I = os.path.join(DATA, 'session.json')
FILE_O = os.path.join(DATA, 'cables.xlsx')

# ==============================================================================
# Data
# ==============================================================================

session = compas.json_load(FILE_I)

mesh = session['mesh']
export = session['export']

cap08 = 0
cap10 = 0

# ==============================================================================
# Workbook
# ==============================================================================

headerfont = Font(bold=True, color="ff0000")

wb = Workbook()
ws = wb.active
ws.title = "Summary"

a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

a1.value = "Type"
b1.value = "Segments"
c1.value = "Total Length [m]"

a1.font = headerfont
b1.font = headerfont
c1.font = headerfont

for size in export:
    cables = export[size]
    cables = sorted(cables, key=lambda x: x['Label'])

    ws_name = "cables {}".format(size)

    ws = wb.create_sheet(title=ws_name)

    ws['A1'] = "Label"
    ws['B1'] = "Code A"
    ws['C1'] = "Code B"
    ws['D1'] = "Confection Length [mm]"

    ws['A1'].font = headerfont
    ws['B1'].font = headerfont
    ws['C1'].font = headerfont
    ws['D1'].font = headerfont

    if size == '6':
        cap08 += 2 * len(cables)
    elif size == '8':
        cap10 += 2 * len(cables)

    for index, line in enumerate(cables):
        row = index + 2

        ws['A{}'.format(row)] = line['Label']
        ws['B{}'.format(row)] = line['Code A']
        ws['C{}'.format(row)] = line['Code B']
        ws['D{}'.format(row)] = round(line['Confection Length'], 1)

    ws['A{}'.format(row + 1)] = "=COUNTA(A2:A{})".format(row)
    ws['D{}'.format(row + 1)] = "=SUM(D2:D{})".format(row)

    wb['Summary'].append([
        ws_name,
        '={}!A{}'.format(utils.quote_sheetname(ws_name), row + 1),
        '={}!D{} * 0.001'.format(utils.quote_sheetname(ws_name), row + 1)])

# ==============================================================================
# Save
# ==============================================================================

wb.save(FILE_O)
